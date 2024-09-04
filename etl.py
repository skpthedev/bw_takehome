"""
ETL (Extract, Transform, Load) module for processing child care provider data.

This module contains functions to clean, transform, and load data from Excel files
into a SQLite database for child care providers.
"""

import os
import re
import sqlite3
import hashlib
from datetime import datetime
from typing import Tuple, Optional, Dict, Any

import numpy as np
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
load_dotenv('.env')

# Constants
DATABASE_NAME = 'child_care_data.db'
EXCEL_FILE_NAME = 'Technical Exercise Data.xlsx'
API_KEY = os.getenv('api_key')

# Database setup
conn = sqlite3.connect(DATABASE_NAME)
cursor = conn.cursor()

# Create the table (updated schema)
cursor.execute('''
CREATE TABLE IF NOT EXISTS child_care_providers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    accepts_financial_aid TEXT,
    ages_served TEXT,
    capacity INTEGER,
    certificate_expiration_date DATE,
    city TEXT,
    address1 TEXT,
    address2 TEXT,
    company TEXT,
    phone TEXT,
    phone2 TEXT,
    county TEXT,
    curriculum_type TEXT,
    email TEXT,
    language TEXT,
    license_status TEXT,
    license_issued DATE,
    license_number INTEGER,
    license_renewed DATE,
    license_type TEXT,
    contact_name TEXT,
    max_age INTEGER,
    min_age INTEGER,
    operator TEXT,
    schedule TEXT,
    state TEXT,
    title TEXT,
    website_address TEXT,
    zip TEXT,
    facility_type TEXT,
    source_file TEXT,
    etl_timestamp DATETIME,
    record_hash TEXT
)
''')


def all_none(data: Dict[str, Any]) -> bool:
    """Check if all values in a dictionary are None."""
    return all(value is None for value in data.values())


def clean_phone(phone: str) -> str:
    """Remove non-digit characters from a phone number string."""
    return re.sub(r'\D', '', str(phone))


def parse_date(date_str: Optional[str]) -> Optional[datetime.date]:
    """Parse a date string into a datetime.date object."""
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    try:
        return datetime.strptime(str(date_str), '%m/%d/%y').date()
    except (ValueError, TypeError):
        return None


def get_ages_served(row: pd.Series) -> Tuple[str, Optional[int], Optional[int]]:
    """Extract ages served information from a data row."""
    age_ranges = {
        'Infants': (0, 11),
        'Toddlers': (12, 23),
        'Preschool': (24, 59),
        'School-age': (60, None)
    }
    
    ages = []
    min_age = None
    max_age = None

    # Convert numpy array to pandas Series if necessary
    if isinstance(row, np.ndarray):
        row = pd.Series(row, index=['Infant', 'Toddler', 'Preschool', 'School'])
    elif isinstance(row, dict):
        row = pd.Series(row)

    # Check if all values are None or NaN
    if row.isnull().all():
        return '', None, None
    
    for age_group, (low, high) in age_ranges.items():
        column_name = age_group[:-1] if age_group.endswith('s') else age_group  # Remove 's' for column matching
        if column_name in row.index and pd.notna(row[column_name]) and str(row[column_name]).upper() == 'Y':
            ages.append(age_group)
            if min_age is None or low < min_age:
                min_age = low
            if max_age is None or (high is not None and high > max_age):
                max_age = high
        elif any(age_group.lower() in str(value).lower() for value in row if pd.notna(value)):
            ages.append(age_group)
            if min_age is None or low < min_age:
                min_age = low
            if max_age is None or (high is not None and high > max_age):
                max_age = high
    
    return ', '.join(ages), min_age, max_age if max_age is not None else None


def extract_title(name: Optional[str]) -> Optional[str]:
    """Extract a title from a name string."""
    title_pattern = r'\b(Director|Owner|Primary Caregiver|Other)\b'
    match = re.search(title_pattern, str(name))
    if match:
        return match.group(1)
    return None


def geocode_address(address: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Geocode an address using the TomTom API."""
    url = f"https://api.tomtom.com/search/2/geocode/{address}.json?key={API_KEY}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if data['results']:
            result = data['results'][0]
            return (
                result['address'].get('countrySubdivision'),
                result['address'].get('municipality'),
                result['address'].get('postalCode')
            )
    return None, None, None


def process_excel_file(file_path: str) -> None:
    """Process the Excel file and load data into the database."""
    wb = load_workbook(filename=file_path, read_only=True)
    
    for sheet_name in ['source1', 'source2', 'source3']:
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            if all_none(row_data):
                continue
            
            # Common data extraction
            age_data = {
                'Ages Accepted 1': row_data.get('Ages Accepted 1'),
                'AA2': row_data.get('AA2'),
                'AA3': row_data.get('AA3'),
                'AA4': row_data.get('AA4'),
                'Infant': row_data.get('Infant'),
                'Toddler': row_data.get('Toddler'),
                'Preschool': row_data.get('Preschool'),
                'School': row_data.get('School')
            }
            
            ages_served, min_age, max_age = get_ages_served(age_data)
            
            contact_name = row_data.get('Primary Contact Name') or row_data.get('Primary Caregiver') or ''
            title = extract_title(contact_name) or row_data.get('Primary Contact Role') or ''
            
            full_address = (
                row_data.get('Address') or
                row_data.get('Address 1') or
                f"{row_data.get('Address')}, {row_data.get('City')} {row_data.get('State')} {row_data.get('Zip')}" or
                None
            )
            
            state, city, zip_code = None, None, None
            if full_address:
                state, city, zip_code = geocode_address(full_address)
            
            data = {
                
                data = {
                    'accepts_financial_aid': str(row_data.get('Accepts Subsidy', '')).lower() == 'accepts subsidy',
                    'ages_served': ages_served or None,
                    'capacity': row_data.get('Total Cap') or row_data.get('Capacity') or None,
                    'certificate_expiration_date': parse_date(row_data.get('Expiration Date')) or None,
                    'city': city or row_data.get('City') or None,
                    'address1': row_data.get('Address') or row_data.get('Address1') or None,
                    'address2': row_data.get('Address2') or None,
                    'company': row_data.get('Name') or row_data.get('Company') or row_data.get('Operation Name') or None,
                    'phone': clean_phone(row_data.get('Phone') or '') or None,
                    'phone2': None,
                    'county': row_data.get('County') or None,
                    'curriculum_type': None,
                    'email': row_data.get('Email') or row_data.get('Email Address') or None,
                    'license_status': row_data.get('Status') or None,
                    'license_issued': parse_date(row_data.get('Issue Date') or row_data.get('First Issue Date')) or None,
                    'license_number': row_data.get('Credential Number') or row_data.get('Operator') or None,
                    'license_renewed': None,
                    'license_type': row_data.get('Credential Type') or row_data.get('Type License') or row_data.get('Type') or None,
                    'contact_name': contact_name or None,
                    'max_age': max_age,
                    'min_age': min_age,
                    'operator': None,
                    'schedule': row_data.get('Year Round', '') or None,
                    'state': state or row_data.get('State') or None,
                    'title': title or None,
                    'website_address': None,
                    'zip': zip_code or row_data.get('Zip') or None,
                    'facility_type': row_data.get('Credential Type') or row_data.get('Type License') or row_data.get('Type') or None,
                    'source_file': sheet_name,
                    'etl_timestamp': datetime.now().isoformat()
                }

                # Generate a hash for the record
                record_hash = hashlib.md5(str(data).encode()).hexdigest()
                data['record_hash'] = record_hash

                # Check if the record already exists
                cursor.execute('SELECT id FROM child_care_providers WHERE (address1 = ? AND city = ? AND state = ?)',
                            (data['address1'], data['city'], data['state']))
                
                existing_record = cursor.fetchone()

                if existing_record:
                    # Update the existing record
                    update_query = 'UPDATE child_care_providers SET ' + ', '.join([f'{k} = ?' for k in data.keys()]) + ' WHERE id = ?'
                    cursor.execute(update_query, list(data.values()) + [existing_record[0]])
                else:
                # Insert a new record
                    insert_query = 'INSERT INTO child_care_providers (' + ', '.join(data.keys()) + ') VALUES (' + ', '.join(['?' if v is not None else 'NULL' for v in data.values()]) + ')'
                    cursor.execute(insert_query, [v for v in data.values() if v is not None])

        conn.commit()

    wb.close()

# Main execution
if __name__ == "__main__":
    process_excel_file('Technical Exercise Data.xlsx')
    conn.close()
    print("ETL process completed successfully.")
